#!/usr/bin/env python3
"""
Extract the 'Catalogue' sheet from CatalogueV22.xlsm and export to JSON
with lowercase column keys.
"""

import pandas as pd
import json
from pathlib import Path
import sys
import re
from openpyxl import load_workbook


def extract_catalogue_to_json(excel_file="Catalogue_V22.xlsm", output_file="catalogue.json"):
    """
    Extract the 'Catalogue' sheet from an Excel file and save as JSON
    with lowercase column keys.

    Args:
        excel_file (str): Path to the Excel file
        output_file (str): Path for the output JSON file
    """
    try:
        # Check if the Excel file exists
        if not Path(excel_file).exists():
            print(f"Error: {excel_file} not found!")
            sys.exit(1)

        print(f"Reading '{excel_file}'...")

        # Load the 'Catalogue' sheet
        df = pd.read_excel(excel_file, sheet_name='Catalogue')

        print(
            f"Found {len(df)} rows and {len(df.columns)} columns in the Catalogue sheet")
        print(f"Original columns: {list(df.columns)}")

        # Extract hyperlinks from DOI column using openpyxl BEFORE any column processing
        print("Extracting hyperlinks from DOI column...")
        wb = load_workbook(excel_file)
        ws = wb['Catalogue']

        # Find DOI column index in original columns
        doi_col_idx = None
        original_columns = list(df.columns)
        for idx, col in enumerate(original_columns):
            if col.upper() == 'DOI':
                doi_col_idx = idx + 1  # openpyxl uses 1-based indexing
                break

        hyperlinks = {}
        if doi_col_idx:
            # Skip header row (row 1), start from row 2
            for row_idx in range(2, len(df) + 2):
                cell = ws.cell(row=row_idx, column=doi_col_idx)
                if cell.hyperlink and cell.hyperlink.target:
                    # Convert to 0-based index
                    hyperlinks[row_idx - 2] = cell.hyperlink.target

        print(f"Found {len(hyperlinks)} hyperlinks in DOI column")

        # Remove columns that start with 'colonne' (case insensitive)
        columns_to_keep = [
            col for col in df.columns if not col.lower().startswith('colonne')]
        df = df[columns_to_keep]

        print(
            f"After removing 'colonne' columns: {len(df.columns)} columns remaining")

        # Convert column names to lowercase
        df.columns = df.columns.str.lower()

        # Fix the typo from 'mounth' to 'month'
        df.columns = df.columns.str.replace('doi', 'url')

        # Fix the typo from 'mounth' to 'month'
        df.columns = df.columns.str.replace('mounth', 'month')
        df.columns = df.columns.str.replace(
            'detailled_parameters', 'p02_parameters')

        print(f"Final columns: {list(df.columns)}")

        # Convert DataFrame to dictionary with records orientation
        data = df.to_dict('records')

        # Handle NaN values and process records
        processed_data = []
        filtered_count = 0
        url_counts = {}  # Track URL usage for duplicate detection
        no_url_records = []  # Track records without URLs
        duplicate_url_records = []  # Track records with duplicate URLs

        for idx, record in enumerate(data):
            # Convert NaN values to None
            for key, value in record.items():
                if pd.isna(value):
                    record[key] = None

            # Check if both name and url are empty/null
            name = record.get('name')
            url = record.get('url')

            # Skip records where both name and url are empty/null
            if (name is None or (isinstance(name, str) and name.strip() == '')) and \
               (url is None or (isinstance(url, str) and url.strip() == '')):
                filtered_count += 1
                print(
                    f"  Filtering out record ID {record.get('id_dataset')}: both name and url are empty")
                continue

            # Replace DOI with hyperlink URL if available
            if idx in hyperlinks:
                record['url'] = hyperlinks[idx]
                print(
                    f"  Updated URL for ID {record.get('id_dataset')}: {hyperlinks[idx]}")

            # Process p02_parameters field - split comma-separated values into a list with lowercase
            p02_params = record.get('p02_parameters')
            if p02_params and isinstance(p02_params, str):
                # Split by comma, strip whitespace, and convert to lowercase
                record['p02_parameters'] = [param.strip().lower()
                                            for param in p02_params.split(',') if param.strip()]
            elif p02_params is None:
                record['p02_parameters'] = []

            # Create bbox field as GeoJSON specification [west, south, east, north]
            bbox = []
            for coord_key in ['bounding_box_w', 'bounding_box_s', 'bounding_box_e', 'bounding_box_n']:
                coord_value = record.get(coord_key)
                if coord_value is not None:
                    try:
                        # Handle comma as decimal separator
                        if isinstance(coord_value, str):
                            coord_value = coord_value.replace(',', '.')
                        bbox.append(float(coord_value))
                    except (ValueError, TypeError):
                        bbox.append(None)
                else:
                    bbox.append(None)

            # Add the bbox field and remove individual coordinate fields
            record['bbox'] = bbox
            for coord_key in ['bounding_box_s', 'bounding_box_n', 'bounding_box_w', 'bounding_box_e']:
                record.pop(coord_key, None)

            # Ensure format field is always a list
            format_value = record.get('format')
            if format_value:
                if isinstance(format_value, str):
                    # Split by comma if there are multiple formats, otherwise make it a single-item list
                    if ',' in format_value:
                        record['format'] = [fmt.strip()
                                            for fmt in format_value.split(',') if fmt.strip()]
                    else:
                        record['format'] = [format_value.strip()]
                elif not isinstance(format_value, list):
                    record['format'] = [str(format_value)]
            else:
                record['format'] = []

            # Rename geographical_coverage_name to basins and make it a lowercase list
            geographical_coverage = record.get('geographical_coverage_name')
            if geographical_coverage and isinstance(geographical_coverage, str):
                # Split by comma and convert to lowercase
                record['regions'] = [basin.strip().lower()
                                     for basin in geographical_coverage.split(',') if basin.strip()]
            elif geographical_coverage is None:
                record['regions'] = []
            else:
                record['regions'] = [str(geographical_coverage).lower()]

            # Remove the old key
            record.pop('geographical_coverage_name', None)

            # Track URL usage for analysis
            url = record.get('url')
            record_id = record.get('id_dataset')

            if not url or (isinstance(url, str) and url.strip() == ''):
                no_url_records.append(record_id)
            else:
                url_str = str(url).strip()
                if url_str in url_counts:
                    url_counts[url_str].append(record_id)
                    if len(url_counts[url_str]) == 2:  # First duplicate
                        duplicate_url_records.extend(url_counts[url_str])
                    else:  # Additional duplicates
                        duplicate_url_records.append(record_id)
                else:
                    url_counts[url_str] = [record_id]

            processed_data.append(record)

        data = processed_data
        print(f"Filtered out {filtered_count} records with empty name and url")

        # Report URL analysis
        print(f"\nURL Analysis:")
        print(f"Records without URLs: {len(no_url_records)}")
        if no_url_records:
            print(f"  IDs: {sorted(no_url_records)}")

        print(
            f"Records with duplicate URLs: {len(set(duplicate_url_records))}")
        if duplicate_url_records:
            print(f"  IDs: {sorted(set(duplicate_url_records))}")

        # Show which URLs are duplicated
        duplicate_urls = {url: ids for url,
                          ids in url_counts.items() if len(ids) > 1}
        if duplicate_urls:
            print(f"Duplicate URL details:")
            for url, ids in sorted(duplicate_urls.items()):
                print(f"  '{url}' used by IDs: {sorted(ids)}")

        # Write to JSON file
        print(f"\nWriting to '{output_file}'...")
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        print(f"Successfully exported {len(data)} records to '{output_file}'")

    except FileNotFoundError:
        print(f"Error: Sheet 'Catalogue' not found in {excel_file}")
        sys.exit(1)
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)


if __name__ == "__main__":
    # Allow command line arguments for custom file paths
    excel_file = sys.argv[1] if len(sys.argv) > 1 else "Catalogue_V22.xlsm"
    output_file = sys.argv[2] if len(sys.argv) > 2 else "catalogue.json"

    extract_catalogue_to_json(excel_file, output_file)
